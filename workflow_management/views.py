
from django.http import HttpResponse

# Create your views here.



from django.shortcuts import render,redirect, get_object_or_404
from django.views import View
from .models import *
from django.views.generic import TemplateView
from django.contrib.auth.models import User
from django.http import JsonResponse
import os, mimetypes, json
from django.contrib.auth.models import User, Group
from django.core.mail import send_mail
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl.styles import Font
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import pandas as pd
import email, imaplib
from email.header import decode_header
# from .utils import fetch_sent_emails
import uuid
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required









def workflowmgt(request):
    return render(request, 'components.html')  

def workflow(request):
    return render(request,"workflow.html")




from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db.models import Q

from django.db.models import Q
from django.utils import timezone

class LeaveRequestView(View):
    allowed_file_types = ['pdf', 'xls', 'xlsx', 'doc', 'docx', 'ppt', 'pptx', 'jpg', 'jpeg', 'png', 'txt']

    def get(self, request, *args, **kwargs):
        # Your leave requests only (submitted by you)
        user_leave_requests = LeaveRequest.objects.filter(user=request.user).order_by('-created_at')

        # Leave requests pending your approval (not created by you)
        pending_approval_requests = LeaveRequest.objects.filter(
            status="Pending",
            current_level__in=[1, 2, 3]
        ).exclude(user=request.user).filter(
            Q(level1_approvers=request.user, current_level=1, status="Pending") |
            Q(level2_approvers=request.user, current_level=2, status="Pending") |
            Q(level3_approvers=request.user, current_level=3, status="Pending")
        ).distinct().order_by('-created_at')

        # Calculate number of days for each leave request
        for leave_request in user_leave_requests:
            if leave_request.from_date and leave_request.to_date:
                leave_request.no_of_days = (leave_request.to_date - leave_request.from_date).days + 1
            else:
                leave_request.no_of_days = 0

        for leave_request in pending_approval_requests:
            if leave_request.from_date and leave_request.to_date:
                leave_request.no_of_days = (leave_request.to_date - leave_request.from_date).days + 1
            else:
                leave_request.no_of_days = 0

        context = {
            'user_leave_requests': user_leave_requests,
            'pending_approval_requests': pending_approval_requests,
            'level1_users': User.objects.filter(groups__name='Level1'),
            'level2_users': User.objects.filter(groups__name='Level2'),
            'level3_users': User.objects.filter(groups__name='Level3'),  # Only pending requests for your approval
        }

        return render(request, 'leavedetails.html', context)
        
    def post(self, request, *args, **kwargs):
        # Handling file validation
        file = request.FILES.get('file')
        max_file_size = 5 * 1024 * 1024
        if file:
            extension = os.path.splitext(file.name)[1][1:].lower()
            if extension not in self.allowed_file_types:
                return JsonResponse({
                    'status': 'error',
                    'message': f"File type not allowed. Allowed types: {', '.join(self.allowed_file_types)}"
                })
        if file!=None:

            if file.size > max_file_size:
                return JsonResponse({
                    'status': 'error',
                    'message': "File size exceeds the 5 MB limit."
                })

        # Save the leave request
        leave_request = LeaveRequest.objects.create(
            user=request.user,
            leave_type=request.POST.get('LeaveType'),
            from_date=request.POST.get('fromdate'),
            to_date=request.POST.get('todate'),
            session_from=request.POST.get('Session1'),
            session_to=request.POST.get('Session2'),
            reason=request.POST.get('Reason'),
            file=file,
            status='Pending',
            created_at=timezone.now(),
            current_level=1
        )

        # Assign approvers and notify them
        for level, field_name in enumerate(['level1_users[]', 'level2_users[]', 'level3_users[]'], start=1):
            user_ids = request.POST.getlist(field_name)
            users = User.objects.filter(id__in=user_ids)
            getattr(leave_request, f'level{level}_approvers').set(users)
            if level == 1:  # Notify Level 1 approvers
                self.notify_approvers(users, leave_request)

        return JsonResponse({'status': 'success', 'message': 'Leave applied successfully!'})

    def notify_approvers(self, approvers, leave_request):
        subject = f"Leave Request Approval Needed from {leave_request.user.username}"
        message = (
            f"A leave request from {leave_request.user.username} requires your approval.\n\n"
            f"Leave Type: {leave_request.leave_type}\n"
            f"From: {leave_request.from_date}\n"
            f"To: {leave_request.to_date}\n"
            f"Reason: {leave_request.reason}\n"
        )
        email = EmailMessage(
            subject,
            message,
            settings.EMAIL_HOST_USER,
            [approver.email for approver in approvers]
        )
        try:
            email.send(fail_silently=False)
        except Exception as e:
            print(f"Failed to send email: {e}")

@csrf_exempt
def approve_leave(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id)
    approvers, next_level_users = None, None
    
    # Determine current level approvers and next level approvers
    if leave_request.current_level == 1:
        approvers = leave_request.level1_approvers.all()
        next_level_users = leave_request.level2_approvers.all()
    elif leave_request.current_level == 2:
        approvers = leave_request.level2_approvers.all()
        next_level_users = leave_request.level3_approvers.all()
    elif leave_request.current_level == 3:
        approvers = leave_request.level3_approvers.all()

    # Check if the user is authorized to approve at the current level
    if request.user not in approvers:
        return JsonResponse({'status': 'error', 'message': 'You are not authorized to approve this request.'})

    # Update leave request's current level or mark as Approved if at the final level
    if leave_request.current_level < 3:
        leave_request.current_level += 1
        leave_request.status = 'Pending'
        leave_request.save()
        
        # Notify the next level approvers if there are any
        if next_level_users:
            LeaveRequestView().notify_approvers(next_level_users, leave_request)
        # messages.success(request, "Leave request moved to the next approval level.")
    else:
        # Mark as Approved when all levels have approved
        leave_request.status = 'Approved'
        leave_request.save()
        # messages.success(request, "Leave request approved successfully.")

   
    return JsonResponse({'status': 'success', 'message': 'Leave request approved successfully.'})

def reject_leave(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id)

    if request.method == 'POST':
        # Get the rejection reason from the request data (from the JavaScript AJAX POST request)
        data = json.loads(request.body)  # Extracting data sent by the frontend (JSON)
        reason = data.get('Reason')

        if reason:
            # Save the reason and update the leave status to 'Rejected'
            leave_request.rejection_reason = reason  # Assuming 'reason' field exists in the model
            leave_request.status = 'Rejected'
            leave_request.save()

            

            # Add a success message to notify that the leave request was rejected
            messages.success(request, "Leave request rejected successfully.")
            print("leave rejected")
             # Send email notification to the user
            subject = "Your Leave Request has been Rejected"
            message = f"Dear {leave_request.user.username},\n\nYour leave request has been rejected.\nReason: {reason}\n\nThank you."
            recipient = leave_request.user.email

            try:
                send_mail(subject, message, settings.EMAIL_HOST_USER, [recipient])
            except Exception as e:
                print(f"Failed to send email: {e}")
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'Reason is required'})
      # Handle GET request (show the leave request details)

class WithdrawLeaveView(LoginRequiredMixin, View):
    def post(self, request, leave_id):
        print("Withdrawal function called")
        
        # Fetch the leave request based on the leave_id
        leave_request = get_object_or_404(LeaveRequest, id=leave_id, user=request.user)
        print("Processing POST request")

        # Delete the leave request
        leave_request.delete()

        return JsonResponse({'status': 'success', 'message': 'Leave request withdrawn successfully.'})
    
    def get(self, request, leave_id):
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})
    

def gst1(req):
    return render(req,"gst.html")
 
 
 
def download_sample_book(request):
    # Create a workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sample Book"
 
    # Add headers to the sheet
    headers = [
        "Branch", "Customer Name", "Billing Address GSTIN", "Posting Date", "Invoice", "Item Name", "HSN Code", "Place of Supply", "Income Account", "Amount", "Output Tax CGST Rate", "Output Tax CGST Amount", "Output Tax IGST Rate", "Output Tax IGST Amount", "Output Tax SGST Rate", "Output Tax SGST Amount", "Total Tax", "Total Other Charges", "Total", "IRN", "e-Invoice Status", "Service Period", "Reverse Charge"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the headers bold
 
    # Set response headers
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="book.xlsx"'
 
    # Save the workbook to the response
    workbook.save(response)
    return response
 
 
 
def download_sample_portal(request):
    # Create a workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sample Portal"
 
    # Add headers to the sheet
    headers = [
        "Return Period", "Supplier Name", "GSTIN of Supplier", "Invoice Number", "Invoice Date", "Invoice Value", "Place of Supply", "Reverse Charge", "Invoice Type", "Taxable Value", "CGST","SGST", "IGST", "Cess Paid"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the headers bold
 
    # Set response headers
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Portal.xlsx"'
 
    # Save the workbook to the response
    workbook.save(response)
    return response


def validate_and_compare_files(request):
    if request.method == "POST":
        book_file = request.FILES.get('book')
        portal_file = request.FILES.get('portal')

        if not book_file or not portal_file:
            return render(request, 'gst.html', {'error': 'Both files are required.'})

        try:
            book_df = pd.read_excel(book_file)
            portal_df = pd.read_excel(portal_file)
        except Exception as e:
            return render(request, 'gst.html', {'error': 'Error reading files. Please upload valid Excel files.'})

        # Expected columns for validation
        book_columns = ["Branch", "Customer Name", "Billing Address GSTIN", "Posting Date", "Invoice", "Item Name", 
                        "HSN Code", "Place of Supply", "Income Account", "Amount", "Output Tax CGST Rate", 
                        "Output Tax CGST Amount", "Output Tax IGST Rate", "Output Tax IGST Amount", 
                        "Output Tax SGST Rate", "Output Tax SGST Amount", "Total Tax", "Total Other Charges", 
                        "Total", "IRN", "e-Invoice Status", "Service Period", "Reverse Charge"]
        
        portal_columns = ["Return Period", "Supplier Name", "GSTIN of Supplier", "Invoice Number", 
                          "Invoice Date", "Invoice Value", "Place of Supply", "Reverse Charge", 
                          "Invoice Type", "Taxable Value", "CGST", "SGST", "IGST", "Cess Paid"]

        # Check if the columns match the expected format
        if not set(book_columns).issubset(set(book_df.columns)):
            return render(request, 'gst.html', {'error': 'Book file format is incorrect.'})
        if not set(portal_columns).issubset(set(portal_df.columns)):
            return render(request, 'gst.html', {'error': 'Portal file format is incorrect.'})

       # Standardize Reverse Charge Values
        reverse_charge_map = {'n': 'No', 'no': 'No', '0': 'No', 'y': 'Yes', 'yes': 'Yes', '1': 'Yes'}

        # Strip whitespace, convert to lowercase, and map
        book_df['Reverse Charge'] = (
            book_df['Reverse Charge']
            .astype(str)  # Convert to string to handle mixed types
            .str.strip()  # Remove leading/trailing spaces
            .str.lower()  # Convert to lowercase
            .map(reverse_charge_map)
            .fillna('No')  # Default to 'No' for missing/invalid values
        )

        portal_df['Reverse Charge'] = (
            portal_df['Reverse Charge']
            .astype(str)  # Convert to string to handle mixed types
            .str.strip()  # Remove leading/trailing spaces
            .str.lower()  # Convert to lowercase
            .map(reverse_charge_map)
            .fillna('No')  # Default to 'No' for missing/invalid values
        )
        # Convert dates for comparison
        book_df['Posting Date'] = pd.to_datetime(book_df['Posting Date'])
        portal_df['Invoice Date'] = pd.to_datetime(portal_df['Invoice Date'])

        # Identify missing invoices
        book_invoices = set(book_df["Invoice"])
        portal_invoices = set(portal_df["Invoice Number"])
        missing_in_book = portal_invoices - book_invoices
        missing_in_portal = book_invoices - portal_invoices

        # Process each portal invoice
        comparison_results = []
        for _, portal_row in portal_df.iterrows():
            invoice_number = portal_row['Invoice Number']
            if invoice_number in missing_in_book:
                continue

            # Fetch book rows for the same invoice
            book_rows = book_df[book_df['Invoice'] == invoice_number]

            # Check for mismatches in specified fields across all rows
            supplier_name_match = all(
                portal_row['Supplier Name'].lower() == row['Customer Name'].lower()
                for _, row in book_rows.iterrows()
            )
            gstin_match = all(
                portal_row['GSTIN of Supplier'] == row['Billing Address GSTIN']
                for _, row in book_rows.iterrows()
            )
            invoice_date_match = all(
                portal_row['Invoice Date'] == row['Posting Date']
                for _, row in book_rows.iterrows()
            )
            reverse_charge_mismatch = [
                (row['Reverse Charge'], portal_row['Reverse Charge'])
                for _, row in book_rows.iterrows()
                if row['Reverse Charge'] != portal_row['Reverse Charge']
            ]
            reverse_charge_match = len(reverse_charge_mismatch) == 0  # No mismatches if list is empty

            # Aggregate book values
            aggregated_values = book_rows[['Amount', 'Output Tax CGST Amount', 'Output Tax SGST Amount', 'Output Tax IGST Amount', 'Total']].sum()
            book_total_value = round(aggregated_values['Total'], 2)
            portal_invoice_value = round(portal_row['Invoice Value'], 2)

            # Apply tolerance check
            result_row = {
                'Invoice Number': invoice_number,
                'Return Period': portal_row['Invoice Date'].strftime('%b-%Y'),
                'Supplier Name': (
                    portal_row['Supplier Name'] if supplier_name_match 
                    else f"Not Matched"
                ),
                'GSTIN of Supplier': (
                    portal_row['GSTIN of Supplier'] if gstin_match 
                    else f"Not Matched"
                ),
                'Invoice Date': (
                    portal_row['Invoice Date'] if invoice_date_match 
                    else f"Not Matched"
                ),
                'Invoice Value': (
                    portal_row['Invoice Value'] if abs(book_total_value - portal_invoice_value) <= 1 
                    else f"Not Matched (Portal: {portal_invoice_value}, Book: {book_total_value})"
                ),
                'Taxable Value': (
                    portal_row['Taxable Value'] 
                    if abs(round(portal_row['Taxable Value'], 2) - round(aggregated_values['Amount'], 2)) <= 1 
                    else f"Not Matched (Portal: {round(portal_row['Taxable Value'], 2)}, Book: {round(aggregated_values['Amount'], 2)})"
                ),
                'CGST': (
                    portal_row['CGST'] 
                    if abs(round(portal_row['CGST'], 2) - round(aggregated_values['Output Tax CGST Amount'], 2)) <= 1 
                    else f"Not Matched (Portal: {round(portal_row['CGST'], 2)}, Book: {round(aggregated_values['Output Tax CGST Amount'], 2)})"
                ),
                'SGST': (
                    portal_row['SGST'] 
                    if abs(round(portal_row['SGST'], 2) - round(aggregated_values['Output Tax SGST Amount'], 2)) <= 1 
                    else f"Not Matched (Portal: {round(portal_row['SGST'], 2)}, Book: {round(aggregated_values['Output Tax SGST Amount'], 2)})"
                ),
                'IGST': (
                    portal_row['IGST'] 
                    if abs(round(portal_row['IGST'], 2) - round(aggregated_values['Output Tax IGST Amount'], 2)) <= 1 
                    else f"Not Matched (Portal: {round(portal_row['IGST'], 2)}, Book: {round(aggregated_values['Output Tax IGST Amount'], 2)})"
                ),
                'Reverse Charge': (
                    portal_row['Reverse Charge']
                    if reverse_charge_match 
                    else f"Not Matched"
                ),
            }
            comparison_results.append(result_row)

        # Write results to Excel
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison Results"

        # Check if comparison_results is not empty
        if comparison_results:
            headers = list(comparison_results[0].keys())
            ws.append(headers)

            # Write comparison results
            for row in comparison_results:
                ws.append([row[col] for col in headers])

            # Highlight mismatches
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(comparison_results) + 1), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    if "Not Matched" in str(cell.value):
                        cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            ws.append(["No comparison results found. Please check the input files."])

        # Write missing invoices to the second sheet
        ws2 = wb.create_sheet(title="Missing Invoices")
        ws2.append(["Invoice Number", "Status"])
        for invoice in missing_in_book:
            ws2.append([invoice, "Missing in Book"])
        for invoice in missing_in_portal:
            ws2.append([invoice, "Missing in Portal"])

        wb.save(output)
        output.seek(0)

        response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="Comparison Results.xlsx"'
        return response

    return render(request, 'gst.html')

def send_email_view(request):
    if request.method == "POST":
        to_email = request.POST.get('to_email')
        subject = request.POST.get('subject')
        body = request.POST.get('body')
        sender_user = request.user
        files = request.FILES.getlist('attachments')  # Get multiple files from the form

        try:
            # Validate file sizes (backend validation for 5 MB limit)
            max_file_size = 5 * 1024 * 1024  # 5 MB in bytes
            for file in files:
                if file.size > max_file_size:
                    raise ValidationError(f"File '{file.name}' exceeds the 5 MB limit.")

            # Generate a unique ID for tracking replies
            unique_id = str(uuid.uuid4())
            tracking_subject = f"{subject} [Ref:{unique_id}]"

            # Append sender's email to the body for clarity
            full_body = f"{body}\n\n---\nThis email was sent by {sender_user.email}."

            # Create the email object
            email = EmailMessage(
                subject=tracking_subject,
                body=full_body,
                from_email='taskaccsys@gmail.com',
                to=[to_email],
                headers={'Reply-To': 'taskaccsys@gmail.com'},  # Replies go to TaskAccSys
            )

            # Attach all files to the email
            for file in files:
                email.attach(file.name, file.read(), file.content_type)

            # Send the email
            email.send(fail_silently=False)

            # Log the email in the database
            sent_email = SentEmail.objects.create(
                message_id=unique_id,
                subject=subject,
                body=body,
                sender=sender_user,
                recipient_email=to_email,
            )

            # Save the attachments in the database
            for file in files:
                Attachment.objects.create(email=sent_email, file=file)

            messages.success(request, "Email sent successfully!")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Error sending email: {str(e)}")

    # Fetch all sent emails by the logged-in user
    sent_emails = SentEmail.objects.filter(sender=request.user).prefetch_related('attachments')
    return render(request, 'email_form.html', {'sent_emails': sent_emails})


import email, imaplib
from email.header import decode_header
from django.conf import settings
from workflow_management.models import SentEmail


def fetch_sent_emails(user_email):
    # Connect to the server
    print("to",user_email)
    mail = imaplib.IMAP4_SSL('imap.gmail.com')  # Use settings from your provider
    mail.login(settings.IMAP_USER, settings.IMAP_PASSWORD)
    
    # Select the "Sent Mail" folder
    mail.select('"[Gmail]/Sent Mail"')  # This is Gmail's "Sent Mail"; adjust for other providers
    
    # Search for emails sent to the specific user
    status, messages = mail.search(None, f'TO "{user_email}"')
    
    # Fetch the emails
    email_ids = messages[0].split()
    emails = []
    
    for email_id in email_ids:
        print("from",email_id)
        # Fetch email by ID
        res, msg_data = mail.fetch(email_id, '(RFC822)')
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                # Parse email content
                msg = email.message_from_bytes(response_part[1])
                subject, encoding = decode_header(msg["Subject"])[0]
                if isinstance(subject, bytes):
                    subject = subject.decode(encoding if encoding else "utf-8")
                from_ = msg.get("From")
                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        if content_type == "text/plain" and "attachment" not in str(part.get("Content-Disposition")):
                            body = part.get_payload(decode=True).decode()
                            break
                else:
                    body = msg.get_payload(decode=True).decode()
                
                # Append to emails list
                emails.append({
                    "from": from_,
                    "subject": subject,
                    "body": body,
                    "date": msg.get("Date"),
                })
    # Logout
    mail.logout()
    return emails

def fetch_and_forward_replies():
    """
    Fetch replies from Gmail inbox and forward them to the original sender, including attachments.
    """
    mail = imaplib.IMAP4_SSL('imap.gmail.com')  # Connect to Gmail
    mail.login(settings.IMAP_USER, settings.IMAP_PASSWORD)  # Login credentials
    mail.select("inbox")  # Select the inbox folder

    # Search for all emails
    status, messages = mail.search(None, "ALL")
    email_ids = messages[0].split()

    for email_id in email_ids:
        # Fetch the email by ID
        res, msg_data = mail.fetch(email_id, "(RFC822)")
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                msg = email.message_from_bytes(response_part[1])

                # Extract the subject and check for the unique tracking ID
                subject = msg.get("Subject")
                message_id = msg.get("Message-ID")  # Unique ID for the reply

                if subject and "[Ref:" in subject and message_id:
                    ref_id = subject.split("[Ref:")[1].split("]")[0]
                    original_email = SentEmail.objects.filter(message_id=ref_id).first()

                    if original_email:
                        # Check if this reply has already been forwarded
                        if ForwardedReply.objects.filter(reply_message_id=message_id).exists():
                            print(f"Reply with Message-ID {message_id} already forwarded.")
                            continue

                        # Extract the reply body and attachments
                        body = ""
                        attachments = []  # List to store attachments
                        if msg.is_multipart():
                            for part in msg.walk():
                                content_type = part.get_content_type()
                                content_disposition = str(part.get("Content-Disposition"))

                                # Handle text/plain content
                                if content_type == "text/plain" and "attachment" not in content_disposition:
                                    body = part.get_payload(decode=True).decode()
                                
                                # Handle attachments
                                elif "attachment" in content_disposition:
                                    file_name = part.get_filename()
                                    if file_name:
                                        attachments.append({
                                            "file_name": file_name,
                                            "file_content": part.get_payload(decode=True),
                                            "content_type": part.get_content_type(),
                                        })
                        else:
                            body = msg.get_payload(decode=True).decode()

                        # Forward the reply to the original sender
                        email_message = EmailMessage(
                            subject=f"Re: {original_email.subject}",
                            body=f"Reply from {msg.get('From')}:\n\n{body}",
                            from_email='taskaccsys@gmail.com',
                            to=[original_email.sender.email],  # Forward to the sender
                        )

                        # Attach files to the forwarded email
                        for attachment in attachments:
                            email_message.attach(
                                attachment["file_name"],
                                attachment["file_content"],
                                attachment["content_type"]
                            )

                        # Send the email
                        email_message.send(fail_silently=False)
                        print(f"Reply forwarded to {original_email.sender.email} with attachments")

                        # Log this reply as forwarded
                        ForwardedReply.objects.create(
                            sent_email=original_email,
                            reply_message_id=message_id,
                        )

    mail.logout()


def fetch_emails_for_user(user_email):
    """
    Fetch all emails from the taskaccsys@gmail.com Sent Mail folder and filter for the user.
    Determine the "From" column dynamically for direct and reply emails.
    """
    # Connect to Gmail
    mail = imaplib.IMAP4_SSL('imap.gmail.com')
    mail.login(settings.IMAP_USER, settings.IMAP_PASSWORD)
    mail.select('"[Gmail]/Sent Mail"')  # Select the "Sent Mail" folder

    # Search for emails sent to the specific user
    status, messages = mail.search(None, f'TO "{user_email}"')
    email_ids = messages[0].split()

    user_emails = []

    for email_id in email_ids:
        # Fetch the email
        res, msg_data = mail.fetch(email_id, "(RFC822)")
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                # Parse the email content
                msg = email.message_from_bytes(response_part[1])
                subject, encoding = decode_header(msg["Subject"])[0]
                if isinstance(subject, bytes):
                    subject = subject.decode(encoding if encoding else "utf-8")
                from_ = "taskaccsys@gmail.com"  # Default to taskaccsys@gmail.com
                to_ = msg.get("To")
                date = msg.get("Date")
                body = ""
                attachments = []

                # Extract the email body
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        content_disposition = str(part.get("Content-Disposition"))

                        # Extract plain text body
                        if content_type == "text/plain" and "attachment" not in content_disposition:
                            body = part.get_payload(decode=True).decode()

                        # Extract attachments
                        # Handle attachments
                        elif "attachment" in content_disposition:
                            file_name = part.get_filename()
                            if file_name:
                                # Save attachment to MEDIA_ROOT
                                sanitized_file_name = file_name.replace(" ", "_")
                                file_path = os.path.join(settings.MEDIA_ROOT, sanitized_file_name)
                                with open(file_path, "wb") as f:
                                    f.write(part.get_payload(decode=True))

                                # Append attachment data with its media URL
                                attachments.append({
                                    "file_name": file_name,
                                    "file_url": f"{settings.MEDIA_URL}{sanitized_file_name}",
                                })
                else:
                    body = msg.get_payload(decode=True).decode()

                # Determine the "From" field
                
                if "Re:" in subject:
                    reply_line = body.split("Reply from")[1].split(":")[0].strip()
                        # Extract the email address within angle brackets < >
                    if "<" in reply_line and ">" in reply_line:
                        from_ = reply_line.split("<")[1].split(">")[0].strip()
                    else:
                        from_ = reply_line  # Fallback if angle brackets are missing

                elif "This email was sent by" in body:
                    # Direct email: Extract the sender's email from the body
                    from_ = body.split("This email was sent by")[1].split(".")[0].strip()+".com"

                # Append the email details to the list
                user_emails.append({
                    "from": from_,  # Resolved "From" field
                    "to": to_,
                    "subject": subject,
                    "body": body,
                    "date": date,
                    "attachments": attachments,
                })

    mail.logout()
    return user_emails



# User Inbox View
def user_inbox_view(request):
    user = request.user
    user_email = user.email  # Get the logged-in user's email

    # Fetch emails sent to the logged-in user
    inbox_emails = fetch_emails_for_user(user_email)

    return render(request, 'inbox.html', {'inbox_emails': inbox_emails})


def send_email_with_tracking(subject, body, recipient_email, sender_user):
    unique_id = str(uuid.uuid4())  # Generate a unique ID
    tracking_subject = f"{subject} [Ref:{unique_id}]"  # Include tracking ID in the subject

    email = EmailMessage(
        subject=tracking_subject,
        body=body,
        from_email='taskaccsys@gmail.com',  # Always sent from the project email
        to=[recipient_email],
        headers={'Reply-To': sender_user.email},  # Include User A's email as Reply-To
    )
    email.send(fail_silently=False)
    return unique_id


def send_and_log_email(subject, body, recipient_email, sender_user):
    unique_id = send_email_with_tracking(subject, body, recipient_email, sender_user)
    SentEmail.objects.create(
        message_id=unique_id,
        subject=subject,
        body=body,
        sender=sender_user,
        recipient_email=recipient_email,
    )

def trigger_fetch_replies(request):
    fetch_and_forward_replies()
    return JsonResponse({'status': 'Replies fetched and forwarded'})

@login_required
def send_reply(request):
    """
    Send a reply email with the user's name and email automatically added to the body.
    """
    if request.method == "POST":
        to_email = request.POST.get('to_email')  # Recipient email
        subject = request.POST.get('subject')  # Subject of the reply
        original_body = request.POST.get('body')  # Original email content

        # Get the logged-in user's name and email
        sender_name = request.user.username or "Anonymous User"
        sender_email = request.user.email

        # Format the reply body
        reply_body = f"Reply from {sender_name} <{sender_email}>:\n\n{original_body}"

        try:
            # Send the reply email
            send_mail(
                subject=subject,
                message=reply_body,
                from_email='taskaccsys@gmail.com',  # Your project email
                recipient_list=[to_email],
                fail_silently=False,
            )
            messages.success(request, "Reply sent successfully!")
        except Exception as e:
            messages.error(request, f"Error sending reply: {str(e)}")
    return redirect('send_email')  # Redirect to the inbox view

